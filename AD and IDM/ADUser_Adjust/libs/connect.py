#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pymysql
from libs.logger import errlog_sys, debuglog_sys
from openpyxl import Workbook, load_workbook
from ldap3 import Server, Connection, SUBTREE
from zeep import Client
from zeep.plugins import HistoryPlugin


class MysqlOper:
    def __init__(self):
        self.host = '10.4.64.2'
        self.user = 'aduser'
        self.passwd = 'ADuser@123'
        self.dbname = 'ad'
        self.charset = 'utf8'
        self.dbconn = self.dbconnect()

        if self.dbconn:
            self.dbcursor = self.dbconn.cursor()

    def dbconnect(self):
        mysqlconn = ''
        try:
            mysqlconn = pymysql.connect(host=self.host,
                                        user=self.user,
                                        passwd=self.passwd,
                                        db=self.dbname,
                                        charset=self.charset)
        except Exception as e:
            errlog_sys(repr(e))
        return mysqlconn

    def dbclose(self):
        if self.dbconn:
            try:
                if self.dbcursor:
                    self.dbcursor.close()
                if self.dbconn:
                    self.dbconn.close()
            except Exception as e:
                errlog_sys(e)

    def dbmanyquery(self, sql, *args):
        """
        执行多行查询语句
        """
        resmany = ''
        if self.dbconn:
            try:
                self.dbcursor.execute(sql, args)
                resmany = self.dbcursor.fetchall()
            except Exception as e:
                errlog_sys(repr(e))
        return resmany

    def dbonequery(self, sql, *args):
        resone = ''
        if self.dbconn:
            try:
                self.dbcursor.execute(sql, args)
                resone = self.dbcursor.fetchone()
            except Exception as e:
                errlog_sys(e)
        return resone

    def dbmanyinsert(self, sql, l_value):
        if self.dbconn:
            try:
                self.dbcursor.executemany(sql, l_value)
                self.dbconn.commit()
                return 1
            except Exception as e:
                self.dbconn.rollback()
                errlog_sys(e)
                return 0

    def dbonemod(self, sql, *args):
        if self.dbconn:
            try:
                self.dbcursor.execute(sql, args)
                self.dbconn.commit()
                return 1
            except Exception as e:
                self.dbconn.rollback()
                errlog_sys(e)
                return 0


class AdOper:
    def __init__(self):
        try:
            # server = Server('192.168.4.100')
            # self.adconn = Connection(server, 'testing\\admin_adsrv', 'Sunac2020', auto_bind=True)
            server = Server('192.168.4.200')
            self.adconn = Connection(server, 'sunac\\jt_srv_ad', 'pass@word1', auto_bind=True)
        except Exception as e:
            errlog_sys(self.adconn.result)

    def adclose(self):
        if self.adconn:
            self.adconn.unbind()

    def adquery(self, searchou, searchfilter):
        if self.adconn:
            try:
                resadquery = self.adconn.search(search_base=searchou,
                                                search_filter=searchfilter)
            except Exception as e:
                errlog_sys(e)
        return resadquery

    def adattrquery(self, searchou, searchfilter, *args):
        if self.adconn:
            try:
                resadattrquery = self.adconn.search(search_base=searchou,
                                                search_filter=searchfilter,
                                                attributes=list(args))
            except Exception as e:
                errlog_sys(e)
        return resadattrquery

    def adadd(self, objdn, objtype):
        if self.adconn:
            try:
                resadadd = self.adconn.add(objdn, objtype)
            except Exception as e:
                errlog_sys(e)
        return resadadd

    def adrename(self, objaddn, objdnname):
        if self.adconn:
            try:
                resadrename = self.adconn.modify_dn(objaddn, objdnname)
            except Exception as e:
                errlog_sys(e)
        return resadrename

    def admove(self, objaddn, objdnname, objnewdn):
        if self.adconn:
            try:
                resadmove = self.adconn.modify_dn(objaddn, objdnname, new_superior=objnewdn)
            except Exception as e:
                errlog_sys(e)
        return resadmove


def idmquery(wsdlurl, begintime, endtime, pagenum, systemid, idmid, clientsrv):
    try:
        history = HistoryPlugin()
        client = Client(wsdl=wsdlurl, plugins=[history])
        querydto_type = client.get_type('ns0:queryDto')
        header_type = client.get_type('ns2:Header')
        querydto = querydto_type(beginDate=begintime, endDate=endtime,
                                 pageNo=pagenum, pageRowNo='100', systemID=systemid)
        header = header_type(BIZTRANSACTIONID=idmid, COUNT='', CONSUMER='',
                             SRVLEVEL='', ACCOUNT='idmadmin', PASSWORD='idmpass')
        if clientsrv == 'idmorg':
            residmquery = client.service.PUBLIC_SUNAC_301_queryIdmOrgData(queryDto=querydto,
                                                                          _soapheaders={'parameters2': header})
        if clientsrv == 'idmuser':
            residmquery = client.service.PUBLIC_SUNAC_300_queryIdmUserData(queryDto=querydto,
                                                                           _soapheaders={'parameters2': header})

        debuglog_sys(history.last_sent)
        debuglog_sys(history.last_received)
    except Exception as e:
        errlog_sys(e)
    return residmquery


def excelidm(excelpath):
    try:
        # 读取excel文档
        wb = load_workbook(excelpath)
        ws = wb['sheet']
        row_totle = ws.max_row
        col_totle = ws.max_column

        l_excelvalue = []
        for rowcell in ws.iter_rows(min_row=2, max_row=row_totle):
            l_rowvalue = []
            for i in range(1, col_totle):
                colvalue = rowcell[i].value
                l_rowvalue.append(colvalue)
            t_rowvalue = tuple(l_rowvalue)
            l_excelvalue.append(t_rowvalue)
    except Exception as e:
        errlog_sys(e)
    return l_excelvalue
