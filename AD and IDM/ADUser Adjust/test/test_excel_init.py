#!/usr/bin/env python
# -*- coding: utf-8 -*-

from libs.connect import MysqlOper, excelidm
from openpyxl import Workbook, load_workbook
import pymysql

EXCELFILE = r"C:\Users\zxcvb\Documents\CloudStation\My Python\SunacProject\AD\file\tmp001-org0106.xlsx"

wb = load_workbook(EXCELFILE)
ws = wb['Sheet1']
row_totle = ws.max_row

dbconn = pymysql.connect(
    host='10.4.64.2',
    user='aduser',
    passwd='ADuser@123',
    db='ad',
    charset='utf8')
dbcursor = dbconn.cursor()

l_orgdepinf = []
for obj_orginfo in ws.iter_rows(min_row=2, max_row=row_totle):
    organnumber = obj_orginfo[1].value
    organname = obj_orginfo[2].value
    organparentno = obj_orginfo[3].value
    organupdate = obj_orginfo[4].value
    organcreate = obj_orginfo[5].value
    organdep = obj_orginfo[6].value
    organstatus = obj_orginfo[7].value

    t_orginf = (organnumber, organname, organparentno, organupdate, organcreate,
                organdep, organstatus)

    l_orgdepinf.append(t_orginf)
print(l_orgdepinf)

insertsql = 'insert into idm_org_init(organnumber, organname, organparentno, organupdate, organcreate, organdep, organstatus) values (%s,%s,%s,%s,%s,%s,%s)'

dbcursor.executemany(insertsql, l_orgdepinf)
dbconn.commit()

dbconn.close()
