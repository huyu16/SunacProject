import pymysql
from openpyxl import Workbook, load_workbook
import re

FILEPATH = r"C:\Users\zxcvb\Documents\CloudStation\tmp001-usr0106.xlsx"
# 读取excel文档
wb = load_workbook(FILEPATH)
ws = wb['sheet']
row_totle = ws.max_row

# 绑定Mysql数据库
dbconn = pymysql.connect(
    host='10.4.64.2',
    user='aduser',
    passwd='ADuser@123',
    db='test_ad',
    charset='utf8')
dbcursor = dbconn.cursor()

l_userinfo = []
# REC = re.compile('OU=融创集团,DC=testing,DC=local')

try:
    for obj_userinfo in ws.iter_rows(min_row=2, max_row=row_totle):
        l_userdepinf = []
        userid = obj_userinfo[1].value
        username = obj_userinfo[2].value
        userdeptno = obj_userinfo[3].value

        if userdeptno != '1' or userdeptno is not None:
            dbcursor.execute('select count(0) from dic_idm_org where organnumber = "%s" '
                             'and orghandled <> 4' % userdeptno)
            num = dbcursor.fetchone()[0]
            if num > 0:
                dbcursor.execute('select orgdep from dic_idm_org where organnumber = "%s" '
                                 'and orghandled <> 4' % userdeptno)
                str_userorg = dbcursor.fetchone()[0]
                l_userorg = str_userorg.split(',')
                str_orghandle = l_userorg[-6:]
                userorg = (','.join(str_orghandle))
            else:
                userorg = 'no organization'
        else:
            userorg = 'no organization'

        userupdate = obj_userinfo[5].value
        usercreate = obj_userinfo[6].value
        useremptype = obj_userinfo[7].value
        userstatus = obj_userinfo[8].value
        pre_userorg = "init pre user org"

        if userorg is not None and userorg.endswith('OU=融创集团,DC=testing,DC=local') and userstatus == 'Active' and useremptype == 'Full-Time':
            userhandled = 11
        else:
            userhandled = 0

        t_userobj = (userid, username, userdeptno, userorg, userupdate,
                     usercreate, useremptype, userstatus, pre_userorg, userhandled)

        l_userinfo.append(t_userobj)

    sql_insert = 'insert into dic_idm_user(userid, username, userdeptno, userorg, userupdate, usercreate, ' \
                 'useremptype, userstatus, pre_userorg, userhandled) ' \
                 'values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'

    dbcursor.executemany(sql_insert, l_userinfo)

    dbconn.commit()
except Exception as e:
    print(userid, username)
    print(e)
