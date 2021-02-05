import pymysql
from openpyxl import Workbook, load_workbook
from ldap3 import Server, Connection, SUBTREE

FILEPATH = r"C:\Users\zxcvb\Documents\CloudStation\tmp002-user1223.xlsx"
# 读取excel文档
wb = load_workbook(FILEPATH)
ws = wb['sheet']
row_totle = ws.max_row

# 绑定Mysql数据库
dbconn = pymysql.connect(
    host='10.4.64.2',
    user='aduser',
    passwd='ADuser@123',
    db='test_ad',
    charset='utf8')
dbcursor = dbconn.cursor()

l_userinfo = []

for obj_userinfo in ws.iter_rows(min_row=2, max_row=row_totle):
    l_userdepinf = []
    userid = obj_userinfo[1].value
    username = obj_userinfo[2].value
    userdeptno = obj_userinfo[3].value

    str_userdep = obj_userinfo[4].value
    l_userdep = str_userdep.split('_')
    userlevel = len(l_userdep)
    for index in range(userlevel):
        if l_userdep[index] == '融创中国':
            l_userdep[index] = '融创集团'
            l_userdepinf.append('OU=' + l_userdep[index])
        else:
            l_userdepinf.append('OU=' + l_userdep[index])
    l_userdepinf.reverse()
    userorg = (','.join(l_userdepinf) + ',DC=testing,DC=local')

    userupdate = obj_userinfo[5].value
    usercreate = obj_userinfo[6].value
    useremptype = obj_userinfo[7].value
    userstatus = obj_userinfo[8].value
    userexpirydate = obj_userinfo[9].value

    if str_userdep is not None and str_userdep.startswith('融创中国') \
            and userstatus == 'Active' and useremptype == 'Full-Time':
        userhandled = 1
    else:
        userhandled = 0

    t_userobj = (userid, username, userdeptno, userorg, userupdate,
                usercreate, useremptype, userstatus, userexpirydate, userhandled)

    l_userinfo.append(t_userobj)

sql_insert = 'insert into dic_idm_user(userid, username, userdeptno, userorg, userupdate, usercreate, ' \
             'useremptype, userstatus, userexpirydate, userhandled) ' \
             'values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'

dbcursor.executemany(sql_insert, l_userinfo)

dbconn.commit()
