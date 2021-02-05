#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pymysql
from openpyxl import Workbook, load_workbook
from ldap3 import Server, Connection, SUBTREE

FILEPATH = r"C:\Users\zxcvb\Documents\CloudStation\My Python\SunacProject\AD\file\tmp001-org0106.xlsx"
# 读取excel文档
wb = load_workbook(FILEPATH)
ws = wb['sheet']
row_totle = ws.max_row

# 绑定Mysql数据库
dbconn = pymysql.connect(
    host='10.4.64.2',
    user='aduser',
    passwd='ADuser@123',
    db='ad',
    charset='utf8')
dbcursor = dbconn.cursor()

server = Server('192.168.4.200')
adconn = Connection(server, 'sunac\\jt_srv_ad', 'pass@word1', auto_bind=True)

l_orginfo = []

for obj_orginfo in ws.iter_rows(min_row=2, max_row=row_totle):
    try:
        l_orgdepinf = []
        organnumber = obj_orginfo[1].value
        organname = obj_orginfo[2].value
        organparentno = obj_orginfo[3].value
        organupdate = obj_orginfo[4].value
        organcreate = obj_orginfo[5].value

        str_orgdep = obj_orginfo[6].value
        l_orgdep = str_orgdep.split('_')
        orglevel = len(l_orgdep)
        for index in range(orglevel):
            if l_orgdep[index] == '融创中国':
                l_orgdep[index] = '融创集团'
                l_orgdepinf.append('OU=' + l_orgdep[index])
            else:
                l_orgdepinf.append('OU=' + l_orgdep[index])
        l_orgdepinf.reverse()
        orgdep = (','.join(l_orgdepinf) + ',DC=SUNAC,DC=local')

        organstatus = obj_orginfo[7].value
        if str_orgdep.startswith('融创中国_融创集团_'):
            pre_orgparentno = "Special preorg num"
            pre_orgdep = "Special preorg"
            orghandled = 4
        elif orglevel > 4:
            pre_orgparentno = "not required preorg num"
            pre_orgdep = "not required preorg"
            orghandled = 6
        else:
            adquery = adconn.search(search_base=orgdep,
                                    search_filter=f'(ou={organname})')

            if adquery:
                pre_orgparentno = "Init Organization Num"
                pre_orgdep = "Init Organization"
                orghandled = 2
            else:
                pre_orgparentno = "no pre Organization Num"
                pre_orgdep = "no pre Organization"
                orghandled = 11

        t_orginf = (organnumber, organname, organparentno, organupdate, organcreate,
                    orgdep, orglevel, organstatus, pre_orgparentno, pre_orgdep, orghandled)
        l_orginfo.append(t_orginf)

    except Exception as e:
        print(orgdep)

    continue

sql_insert = 'insert into dic_idm_org(organnumber, organname, organparentno, organupdate, organcreate, orgdep, orglevel, organstatus, pre_orgparentno, pre_orgdep, orghandled) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'

dbcursor.executemany(sql_insert, l_orginfo)

dbconn.commit()

adconn.unbind()
dbcursor.close()
dbconn.close()
